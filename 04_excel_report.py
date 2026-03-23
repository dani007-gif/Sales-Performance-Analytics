"""
Sales Performance Project
Step 4: Excel Report (openpyxl)
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
import logging
import os

os.makedirs("../logs", exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler("../logs/04_excel_report.log"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ── Helpers ───────────────────────────────────────────────────────────────────
def header_style(ws, row, col_count, title, bg="1a1a2e", fg="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color=fg, size=12, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

def col_headers(ws, row, headers, bg="0f3460", fg="FFFFFF"):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color=fg, size=10, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="medium", color="CCCCCC"))

def auto_width(ws, min_w=10, max_w=30):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        first = col[0]
        if isinstance(first, MergedCell) or not hasattr(first, "column_letter"):
            continue
        values = [str(c.value or "") for c in col if not isinstance(c, MergedCell)]
        if not values:
            continue
        w = max(len(v) for v in values) + 2
        ws.column_dimensions[first.column_letter].width = max(min_w, min(w, max_w))

def data_row(ws, row, values, alt=False):
    bg = "F0F4FF" if alt else "FFFFFF"
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        cell.font = Font(name="Calibri", size=10)

# ── Load aggregated data ──────────────────────────────────────────────────────
df = pd.read_csv("../data/sales_raw.csv", parse_dates=["order_date"])
completed = df[df.status == "Completed"].copy()
completed["year"] = completed.order_date.dt.year
completed["month"] = completed.order_date.dt.to_period("M").astype(str)
completed["quarter"] = "Q" + completed.order_date.dt.quarter.astype(str)

# Aggregations
rev_year = completed.groupby("year").agg(
    orders=("order_id","count"), revenue=("net_revenue","sum"),
    profit=("profit","sum")).reset_index()
rev_region = completed.groupby("region").agg(
    orders=("order_id","count"), revenue=("net_revenue","sum"),
    profit=("profit","sum")).sort_values("revenue",ascending=False).reset_index()
rev_cat = completed.groupby("category").agg(
    orders=("order_id","count"), revenue=("net_revenue","sum"),
    profit=("profit","sum")).sort_values("revenue",ascending=False).reset_index()
rev_rep = completed.groupby("sales_rep").agg(
    orders=("order_id","count"), revenue=("net_revenue","sum"),
    profit=("profit","sum")).sort_values("revenue",ascending=False).reset_index()
rev_channel = completed.groupby("channel").agg(
    orders=("order_id","count"), revenue=("net_revenue","sum"),
    profit=("profit","sum")).sort_values("revenue",ascending=False).reset_index()

# ── Build Workbook ────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ── Sheet 1: Executive Dashboard ─────────────────────────────────────────────
ws = wb.create_sheet("Executive Dashboard")
ws.sheet_view.showGridLines = False

# Title
ws.merge_cells("A1:H1")
ws["A1"] = "SALES PERFORMANCE DASHBOARD 2022–2024"
ws["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
ws["A1"].fill = PatternFill("solid", fgColor="1a1a2e")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# KPI Cards row
kpi_labels = ["Total Revenue", "Total Profit", "Avg Margin", "Total Orders",
               "Completed", "Avg Order Value", "Top Region", "Top Rep"]
kpi_values = [
    f"${completed.net_revenue.sum()/1e6:.1f}M",
    f"${completed.profit.sum()/1e6:.1f}M",
    f"{completed.profit_margin.mean()*100:.1f}%",
    f"{len(completed):,}",
    f"{len(completed[completed.status=='Completed']) if 'status' in completed else len(completed):,}",
    f"${completed.net_revenue.mean():,.0f}",
    completed.groupby("region")["net_revenue"].sum().idxmax(),
    completed.groupby("sales_rep")["net_revenue"].sum().idxmax(),
]
kpi_bg = ["0f3460","533483","e94560","2ec4b6","20bf6b","f7b731","4b7bec","fc5c65"]

ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 30
ws.row_dimensions[5].height = 18

for i, (lbl, val, bg) in enumerate(zip(kpi_labels, kpi_values, kpi_bg), 1):
    col = get_column_letter(i)
    ws[f"{col}3"] = lbl
    ws[f"{col}3"].font = Font(bold=True, color="BBBBBB", size=9, name="Calibri")
    ws[f"{col}3"].fill = PatternFill("solid", fgColor="16213e")
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

    ws[f"{col}4"] = val
    ws[f"{col}4"].font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    ws[f"{col}4"].fill = PatternFill("solid", fgColor=bg)
    ws[f"{col}4"].alignment = Alignment(horizontal="center", vertical="center")

    ws[f"{col}5"] = ""
    ws[f"{col}5"].fill = PatternFill("solid", fgColor=bg)

for i in range(1, 9):
    ws.column_dimensions[get_column_letter(i)].width = 16

# Year summary table
header_style(ws, 8, 5, "Yearly Revenue & Profit Summary")
col_headers(ws, 9, ["Year", "Orders", "Revenue ($)", "Profit ($)", "Margin %"])
for r, row in rev_year.iterrows():
    data_row(ws, 10 + r, [
        row.year,
        row.orders,
        f"${row.revenue:,.0f}",
        f"${row.profit:,.0f}",
        f"{row.profit/row.revenue*100:.1f}%"
    ], alt=r%2==0)

# Region summary table
header_style(ws, 16, 5, "Revenue by Region")
col_headers(ws, 17, ["Region", "Orders", "Revenue ($)", "Profit ($)", "Margin %"])
for r, row in rev_region.iterrows():
    data_row(ws, 18 + r, [
        row.region, row.orders,
        f"${row.revenue:,.0f}", f"${row.profit:,.0f}",
        f"{row.profit/row.revenue*100:.1f}%"
    ], alt=r%2==0)

auto_width(ws)
log.info("Sheet 'Executive Dashboard' created")

# ── Sheet 2: Raw Data ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Raw Data")
cols = list(df.columns)
col_headers(ws2, 1, cols)
for r_idx, row in df.head(500).iterrows():
    for c_idx, val in enumerate(row.values, 1):
        cell = ws2.cell(row=r_idx+2, column=c_idx, value=val)
        cell.font = Font(name="Calibri", size=9)
        if (r_idx % 2) == 0:
            cell.fill = PatternFill("solid", fgColor="F5F5F5")
auto_width(ws2)
log.info("Sheet 'Raw Data' created (500 rows sample)")

# ── Sheet 3: Category Analysis ────────────────────────────────────────────────
ws3 = wb.create_sheet("Category Analysis")
ws3.sheet_view.showGridLines = False
header_style(ws3, 1, 5, "Revenue & Profit by Category")
col_headers(ws3, 2, ["Category", "Orders", "Revenue ($)", "Profit ($)", "Margin %"])
for r, row in rev_cat.iterrows():
    data_row(ws3, 3+r, [
        row.category, row.orders,
        f"${row.revenue:,.0f}", f"${row.profit:,.0f}",
        f"{row.profit/row.revenue*100:.1f}%"
    ], alt=r%2==0)

# Bar chart
chart = BarChart()
chart.type = "col"; chart.title = "Revenue by Category"
chart.y_axis.title = "Revenue ($)"; chart.x_axis.title = "Category"
chart.style = 10; chart.width = 20; chart.height = 12
data_ref = Reference(ws3, min_col=3, max_col=4, min_row=2, max_row=2+len(rev_cat))
cats_ref = Reference(ws3, min_col=1, min_row=3, max_row=2+len(rev_cat))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws3.add_chart(chart, "G2")
auto_width(ws3)
log.info("Sheet 'Category Analysis' created")

# ── Sheet 4: Sales Rep Leaderboard ────────────────────────────────────────────
ws4 = wb.create_sheet("Rep Leaderboard")
ws4.sheet_view.showGridLines = False
header_style(ws4, 1, 5, "Sales Representative Leaderboard")
col_headers(ws4, 2, ["Rank", "Sales Rep", "Orders", "Revenue ($)", "Profit ($)"])
medal = {0:"🥇", 1:"🥈", 2:"🥉"}
for r, row in rev_rep.iterrows():
    prefix = medal.get(r, "")
    data_row(ws4, 3+r, [
        f"{prefix} #{r+1}",
        row.sales_rep, row.orders,
        f"${row.revenue:,.0f}", f"${row.profit:,.0f}"
    ], alt=r%2==0)

# Horizontal bar chart for reps
chart4 = BarChart()
chart4.type = "bar"; chart4.title = "Sales Rep Revenue"
chart4.style = 10; chart4.width = 22; chart4.height = 14
data_ref4 = Reference(ws4, min_col=4, max_col=4, min_row=2, max_row=2+len(rev_rep))
cats_ref4 = Reference(ws4, min_col=2, min_row=3, max_row=2+len(rev_rep))
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref4)
ws4.add_chart(chart4, "G2")
auto_width(ws4)
log.info("Sheet 'Rep Leaderboard' created")

# ── Sheet 5: Channel Analysis ─────────────────────────────────────────────────
ws5 = wb.create_sheet("Channel Analysis")
ws5.sheet_view.showGridLines = False
header_style(ws5, 1, 4, "Performance by Sales Channel")
col_headers(ws5, 2, ["Channel", "Orders", "Revenue ($)", "Profit ($)"])
for r, row in rev_channel.iterrows():
    data_row(ws5, 3+r, [
        row.channel, row.orders,
        f"${row.revenue:,.0f}", f"${row.profit:,.0f}"
    ], alt=r%2==0)
auto_width(ws5)
log.info("Sheet 'Channel Analysis' created")

# ── Save ──────────────────────────────────────────────────────────────────────
os.makedirs("../excel", exist_ok=True)
out = "../excel/Sales_Performance_Report.xlsx"
wb.save(out)
log.info("Excel report saved → %s", out)
print(f"\nExcel report saved: {out}")
